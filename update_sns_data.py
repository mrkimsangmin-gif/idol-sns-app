#!/usr/bin/env python3
"""
update_sns_data.py - SNS 데이터 자동 업데이트 스크립트

Usage:
    C:/Python314/python.exe update_sns_data.py [--force]

    1. 99.input/ 폴더에서 GIRLBAND/BOYBAND 엑셀 자동 감지
    2. 최신 YYMM 시트 자동 감지 후 엑셀 읽기
    3. sns-female.json, sns-male.json 갱신
    4. initial-data.json 재생성 (16조합 Top 30)
    5. 처리된 엑셀을 done/ 폴더로 복사
"""
import sys
import os
import io
import json
import re
import shutil
from datetime import datetime, timezone
from pathlib import Path

# UTF-8 stdout (Windows)
if sys.platform == 'win32' and hasattr(sys.stdout, 'buffer'):
    try:
        sys.stdout = io.TextIOWrapper(sys.stdout.detach(), encoding='utf-8', errors='replace')
        sys.stderr = io.TextIOWrapper(sys.stderr.detach(), encoding='utf-8', errors='replace')
    except Exception:
        pass

try:
    import openpyxl
except ImportError:
    print("openpyxl 필요: pip install openpyxl")
    sys.exit(1)

# =============================================
# 경로 설정
# =============================================
SCRIPT_DIR = Path(__file__).resolve().parent

# idol-sns-app 루트 (직접 지정 — 스크립트가 어디서 실행되든 동일)
APP_DIR = Path('G:/내 드라이브/01.Work/04.AI.M.Contents/00.pumit/04.aimcontents.com/idol-sns-app')
PROJECT_ROOT = Path('G:/내 드라이브/01.Work/04.AI.M.Contents/00.pumit')

INPUT_DIR = PROJECT_ROOT / '99.input'
DONE_DIR = INPUT_DIR / 'done'

DATA_DIR = APP_DIR / 'data'
SNS_FEMALE_JSON = DATA_DIR / 'sns-female.json'
SNS_MALE_JSON = DATA_DIR / 'sns-male.json'
INITIAL_DATA_JSON = DATA_DIR / 'initial-data.json'

# =============================================
# 엑셀 컬럼 -> SNS 플랫폼 매핑
# =============================================
EXCEL_TO_SNS = {
    'weibo': '웨이보',
    'chaohua': '차오화',
    'x_followers': 'X(트위터)',
    'bilibili': '빌리빌리',
    'youtube': '유튜브',
    'qqmusic': 'QQ뮤직',
    'spotify': '스포티파이',
    'instagram': '인스타그램',
}

SNS_LIST = ['웨이보', '차오화', '빌리빌리', 'QQ뮤직', 'X(트위터)', '유튜브', '스포티파이', '인스타그램']

EMPTY_VALUES = {None, '', '.', 0, '0', '-'}


def yymm_to_date(yymm):
    """YYMM -> yyyy-MM (예: 2602 -> 2026-02)"""
    s = str(yymm).strip()
    return f"20{s[:2]}-{s[2:]}"


def detect_excel_files(input_dir):
    """99.input/에서 GIRLBAND/BOYBAND xlsx 자동 감지"""
    girl = sorted(f for f in input_dir.glob('SNS_DATA_GIRLBAND_*-ai.xlsx') if not f.name.startswith('~$'))
    boy = sorted(f for f in input_dir.glob('SNS_DATA_BOYBAND_*-ai.xlsx') if not f.name.startswith('~$'))
    if not girl:
        print("  GIRLBAND 파일 없음")
    if not boy:
        print("  BOYBAND 파일 없음")
    return girl, boy


def detect_latest_sheet(wb):
    """워크북에서 가장 큰 YYMM 시트 찾기"""
    sheets = [n for n in wb.sheetnames if re.match(r'^\d{4}$', n)]
    if not sheets:
        return None
    sheets.sort(key=lambda x: int(x))
    return sheets[-1]


def read_excel_sheet(ws):
    """엑셀 시트 -> 딕셔너리 리스트 (헤더 기반)"""
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    header_idx = None
    for i, row in enumerate(rows):
        for cell in row:
            if cell and str(cell).strip() == 'Name_Korean':
                header_idx = i
                break
        if header_idx is not None:
            break

    if header_idx is None:
        print("    Name_Korean 헤더 없음")
        return []

    headers = [str(h).strip() if h else f'_col_{j}' for j, h in enumerate(rows[header_idx])]
    data = []
    for row in rows[header_idx + 1:]:
        if not any(row):
            continue
        record = {headers[j]: val for j, val in enumerate(row) if j < len(headers)}
        if record.get('Name_Korean'):
            data.append(record)
    return data


def excel_to_records(data, yymm):
    """Wide -> Long 변환 (8 SNS x N groups)"""
    date_str = yymm_to_date(yymm)
    records_by_sns = {sns: [] for sns in SNS_LIST}
    stats = {'total_groups': 0, 'total_records': 0, 'skipped_empty': 0, 'groups': []}

    for row in data:
        name_kr = str(row.get('Name_Korean', '')).strip()
        name_en = str(row.get('Name_English', '')).strip()
        if not name_kr:
            continue

        stats['total_groups'] += 1
        stats['groups'].append(name_kr)

        for excel_col, sns_name in EXCEL_TO_SNS.items():
            val = row.get(excel_col)
            if val in EMPTY_VALUES:
                stats['skipped_empty'] += 1
                continue
            try:
                count = int(float(val))
            except (ValueError, TypeError):
                stats['skipped_empty'] += 1
                continue
            if count <= 0:
                stats['skipped_empty'] += 1
                continue

            records_by_sns[sns_name].append({
                'name': name_kr,
                'group': name_en,
                'date': date_str,
                'count': count,
            })
            stats['total_records'] += 1

    return records_by_sns, stats


def load_json(path):
    """JSON 파일 로드"""
    if not path.exists():
        return None
    with open(str(path), 'r', encoding='utf-8') as f:
        return json.load(f)


def save_json(path, data):
    """JSON 파일 저장 (Google Drive 호환 — 직접 write, tmp+replace 사용하지 않음)"""
    path_str = str(path)

    with open(path_str, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
        f.flush()
        os.fsync(f.fileno())

    size = os.path.getsize(path_str)
    print(f"    저장: {path.name} ({size:,} bytes)")

    # verify
    with open(path_str, 'r', encoding='utf-8') as f:
        verify = json.load(f)
    if 'version' not in verify:
        print(f"    [!] 저장 검증 실패: {path.name}")
        return False
    return True


def merge_sns_data(existing, new_records_by_sns, date_str, gender):
    """기존 sns JSON에 새 월 데이터 병합"""
    if existing is None:
        existing = {
            'generated': '',
            'version': 3,
            'gender': gender,
            'snsList': list(SNS_LIST),
            'data': {},
        }

    # snsList 업데이트 (인스타그램 추가)
    if '인스타그램' not in existing.get('snsList', []):
        existing['snsList'] = list(SNS_LIST)

    dup_platforms = []

    for sns_name, new_records in new_records_by_sns.items():
        if not new_records:
            continue

        if sns_name not in existing['data']:
            existing['data'][sns_name] = {'months': [], 'records': []}

        platform = existing['data'][sns_name]

        # 중복 월 체크
        if date_str in platform.get('months', []):
            dup_platforms.append(sns_name)
            platform['records'] = [r for r in platform['records'] if r['date'] != date_str]
            platform['months'] = [m for m in platform['months'] if m != date_str]

        # 새 레코드 추가
        platform['records'].extend(new_records)
        platform['months'].append(date_str)
        platform['months'] = sorted(set(platform['months']))

        # 레코드 정렬: date desc, count desc
        platform['records'].sort(key=lambda r: (-int(r['date'].replace('-', '')), -r['count']))

    now = datetime.now(timezone.utc).strftime('%Y-%m-%dT%H:%M:%S.000Z')
    existing['generated'] = now
    return existing, dup_platforms


def generate_initial_data(female_data, male_data):
    """initial-data.json 생성 (16조합 Top 30 + 전월)"""
    combinations = {}

    for gender, data in [('여자', female_data), ('남자', male_data)]:
        if data is None:
            continue
        for sns_name in data.get('snsList', []):
            key = f"{gender}_{sns_name}"
            if sns_name not in data.get('data', {}):
                continue

            platform = data['data'][sns_name]
            months = sorted(platform.get('months', []))
            records = platform.get('records', [])
            if not months:
                continue

            latest_month = months[-1]
            prev_month = months[-2] if len(months) >= 2 else None

            latest_recs = sorted([r for r in records if r['date'] == latest_month],
                                 key=lambda r: -r['count'])[:30]
            prev_recs = []
            if prev_month:
                prev_recs = sorted([r for r in records if r['date'] == prev_month],
                                   key=lambda r: -r['count'])[:30]

            combinations[key] = {
                'meta': {
                    'latestMonth': latest_month,
                    'prevMonth': prev_month,
                    'allMonths': months,
                },
                'data': latest_recs + prev_recs,
            }

    now = datetime.now(timezone.utc).strftime('%Y-%m-%dT%H:%M:%S.000Z')
    return {'generated': now, 'version': 2, 'combinations': combinations}


def rebuild_from_all_sheets(excel_path, gender):
    """엑셀의 모든 YYMM 시트를 읽어서 전체 히스토리 재구축"""
    wb = openpyxl.load_workbook(str(excel_path), data_only=True)
    sheets = sorted([n for n in wb.sheetnames if re.match(r'^\d{4}$', n)], key=lambda x: int(x))

    if not sheets:
        print("    YYMM 시트 없음")
        wb.close()
        return None

    print(f"    시트 {len(sheets)}개: {sheets[0]}~{sheets[-1]}")

    result = {
        'generated': '',
        'version': 3,
        'gender': gender,
        'snsList': list(SNS_LIST),
        'data': {},
    }

    total_records = 0
    for sheet_name in sheets:
        ws = wb[sheet_name]
        data = read_excel_sheet(ws)
        if not data:
            print(f"      {sheet_name}: 데이터 없음 (스킵)")
            continue

        records_by_sns, stats = excel_to_records(data, sheet_name)
        date_str = yymm_to_date(sheet_name)
        month_records = 0

        for sns_name, recs in records_by_sns.items():
            if not recs:
                continue
            if sns_name not in result['data']:
                result['data'][sns_name] = {'months': [], 'records': []}
            platform = result['data'][sns_name]
            platform['records'].extend(recs)
            platform['months'].append(date_str)
            platform['months'] = sorted(set(platform['months']))
            month_records += len(recs)

        total_records += month_records
        print(f"      {sheet_name} -> {date_str}: {stats['total_groups']}그룹, {month_records}레코드")

    wb.close()

    # 정렬: date desc, count desc
    for platform in result['data'].values():
        platform['records'].sort(key=lambda r: (-int(r['date'].replace('-', '')), -r['count']))

    now = datetime.now(timezone.utc).strftime('%Y-%m-%dT%H:%M:%S.000Z')
    result['generated'] = now
    print(f"    완료: {len(result['data'])}개 플랫폼, {total_records:,}건")
    return result


def main():
    force = '--force' in sys.argv
    rebuild = '--rebuild' in sys.argv

    print("=" * 60)
    print("  SNS 데이터 자동 업데이트")
    if rebuild:
        print("  [REBUILD 모드: 모든 시트에서 전체 재구축]")
    print("=" * 60)

    # 1. 엑셀 파일 감지
    print(f"\n[1] 입력 폴더: {INPUT_DIR}")
    if not INPUT_DIR.exists():
        print(f"    폴더 없음: {INPUT_DIR}")
        return

    girl_files, boy_files = detect_excel_files(INPUT_DIR)
    if not girl_files and not boy_files:
        print("    처리할 파일 없음")
        return

    DONE_DIR.mkdir(parents=True, exist_ok=True)
    processed_files = []

    # REBUILD 모드: 모든 시트에서 전체 재구축
    if rebuild:
        for gender, files, json_path in [
            ('여자', girl_files, SNS_FEMALE_JSON),
            ('남자', boy_files, SNS_MALE_JSON),
        ]:
            if not files:
                continue
            excel_path = files[0]  # 첫 번째 파일만 사용
            print(f"\n{'─' * 50}")
            print(f"[REBUILD] {excel_path.name} ({gender})")
            result = rebuild_from_all_sheets(excel_path, gender)
            if result:
                save_json(json_path, result)
                processed_files.append((excel_path, gender))

        if processed_files:
            # initial-data.json 재생성
            print(f"\n{'─' * 50}")
            print("[REBUILD] initial-data.json 재생성")
            female_data = load_json(SNS_FEMALE_JSON)
            male_data = load_json(SNS_MALE_JSON)
            initial = generate_initial_data(female_data, male_data)
            save_json(INITIAL_DATA_JSON, initial)
            print(f"     {len(initial['combinations'])}개 조합")

        print(f"\n{'=' * 60}")
        print("  REBUILD 완료!")
        print(f"{'=' * 60}")
        return

    # 2-9. 각 파일 처리 (일반 모드: 최신 시트만)
    for gender, files, json_path in [
        ('여자', girl_files, SNS_FEMALE_JSON),
        ('남자', boy_files, SNS_MALE_JSON),
    ]:
        for excel_path in files:
            print(f"\n{'─' * 50}")
            print(f"[2] {excel_path.name} ({gender})")

            wb = openpyxl.load_workbook(str(excel_path), data_only=True)

            latest_sheet = detect_latest_sheet(wb)
            if not latest_sheet:
                print("    YYMM 시트 없음")
                wb.close()
                continue

            date_str = yymm_to_date(latest_sheet)
            print(f"    시트: {latest_sheet} -> {date_str}")

            ws = wb[latest_sheet]
            data = read_excel_sheet(ws)
            wb.close()

            if not data:
                print("    데이터 없음")
                continue

            # Wide -> Long
            records_by_sns, stats = excel_to_records(data, latest_sheet)

            print(f"    그룹: {stats['total_groups']}개 | 레코드: {stats['total_records']}건 | 스킵: {stats['skipped_empty']}건")
            for sns in SNS_LIST:
                cnt = len(records_by_sns[sns])
                if cnt > 0:
                    print(f"      {sns}: {cnt}")

            # 기존 JSON 로드
            print(f"\n    기존 JSON 로드: {json_path.name}")
            existing = load_json(json_path)
            if existing:
                old_months = set()
                for pdata in existing.get('data', {}).values():
                    old_months.update(pdata.get('months', []))
                print(f"    기존 데이터: {len(existing.get('data',{}))}개 플랫폼, {len(old_months)}개 월")
            else:
                print(f"    기존 파일 없음 - 신규 생성")

            # 중복 월 체크
            if existing and not force:
                dup_check = []
                for sns_name, recs in records_by_sns.items():
                    if not recs:
                        continue
                    if sns_name in existing.get('data', {}):
                        if date_str in existing['data'][sns_name].get('months', []):
                            dup_check.append(sns_name)
                if dup_check:
                    print(f"\n    [!] {date_str} 이미 존재: {', '.join(dup_check)}")
                    answer = input("    덮어쓸까요? (y/N): ").strip().lower()
                    if answer != 'y':
                        print("    건너뜀")
                        continue

            # 병합
            existing, dup_platforms = merge_sns_data(existing, records_by_sns, date_str, gender)
            if dup_platforms:
                print(f"    중복 월 덮어쓰기: {', '.join(dup_platforms)}")

            # 신규 그룹
            all_existing = set()
            for pdata in existing.get('data', {}).values():
                for r in pdata.get('records', []):
                    if r['date'] != date_str:
                        all_existing.add(r['name'])
            new_groups = [g for g in stats['groups'] if g not in all_existing]
            if new_groups:
                print(f"    신규 그룹 {len(new_groups)}개: {', '.join(new_groups[:10])}")
                if len(new_groups) > 10:
                    print(f"      ... 외 {len(new_groups) - 10}개")

            # 병합 후 통계
            total_records = sum(len(p['records']) for p in existing.get('data', {}).values())
            total_months = set()
            for p in existing.get('data', {}).values():
                total_months.update(p.get('months', []))
            print(f"    병합 결과: {len(existing['data'])}개 플랫폼, {len(total_months)}개 월, {total_records:,}건")

            # 저장
            save_json(json_path, existing)
            processed_files.append((excel_path, gender))

    if not processed_files:
        print("\n처리 파일 없음")
        return

    # 10. initial-data.json 재생성
    print(f"\n{'─' * 50}")
    print("[10] initial-data.json 재생성")

    female_data = load_json(SNS_FEMALE_JSON)
    male_data = load_json(SNS_MALE_JSON)
    initial = generate_initial_data(female_data, male_data)
    combo_count = len(initial['combinations'])
    save_json(INITIAL_DATA_JSON, initial)
    print(f"     {combo_count}개 조합")

    for key, combo in sorted(initial['combinations'].items()):
        latest = combo['meta']['latestMonth']
        data_count = len(combo['data'])
        print(f"     {key}: {data_count}건 (최신: {latest})")

    # 11. 처리 파일 복사 (이동이 아닌 복사 — Google Drive 호환)
    print(f"\n{'─' * 50}")
    print("[11] 파일 복사 -> done/")
    for excel_path, gender in processed_files:
        dest = DONE_DIR / excel_path.name
        try:
            shutil.copy2(str(excel_path), str(dest))
            print(f"     복사: {excel_path.name} -> done/")
        except Exception as e:
            print(f"     [!] 복사 실패: {excel_path.name} ({e})")

    # 12. 완료
    print(f"\n{'=' * 60}")
    print("  완료!")
    print(f"  파일: {len(processed_files)}개")
    print(f"  출력: {SNS_FEMALE_JSON.name}, {SNS_MALE_JSON.name}, {INITIAL_DATA_JSON.name}")
    print(f"{'=' * 60}")


if __name__ == '__main__':
    main()
